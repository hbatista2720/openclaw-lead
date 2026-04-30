#!/usr/bin/env python3
"""
Convierte el reporte de llamadas (xlsx) en CSV listo para importar en Twenty CRM.

Uso:
  pip install openpyxl
  python3 scripts/prepare-twenty-import.py \\
    "/ruta/al/archivo.xlsx" \\
    --out infra/twenty-import/empresas_seguimiento.csv

Prioriza la hoja "Hoja 4" (cabecera en fila 1). Si no existe, usa "Abril_Funnel"
saltando la primera fila de título.
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from datetime import date, datetime
from pathlib import Path


def _norm(s: str | None) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).strip())


def _cell(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return _norm(str(v))


def _clean_header(h: str | None) -> str:
    return _norm(h).lower().replace("\xa0", " ")


def _row_to_dict(headers: list[str], row: tuple) -> dict[str, str]:
    out: dict[str, str] = {}
    for i, h in enumerate(headers):
        if not h:
            continue
        key = _clean_header(h)
        val = _cell(row[i]) if i < len(row) else ""
        out[key] = val
    return out


# Claves normalizadas del Excel (cabeceras tras lower + trim)
_HDR_MAP = {
    "nombre empresa": "nombre",
    "nombre de la empresa": "nombre",
    "closer de venta": "closer",
    "etapa de la venta": "etapa",
    "nombre de redes sociales": "red_social",
    "estatus": "estatus",
    "fecha ultima actualizacion": "fecha_actualizacion",
    "servicios activo ?": "servicios",
    "servicios activo?": "servicios",
    "cantidad de lineas movil o servicios": "cantidad_lineas",
    "arpu _ valor del contrato": "valor_contrato",
    "fecha de cierre": "fecha_cierre",
    "contacto telefono": "telefono",
    "recepcion - asistente": "recepcion",
    "encargado principal": "encargado",
    "correo electronico": "correo",
    "observaciones": "observaciones",
}


def _merge_aliases(raw: dict[str, str]) -> dict[str, str]:
    out: dict[str, str] = {}
    for header_key, canon in _HDR_MAP.items():
        v = raw.get(header_key)
        if v:
            out[canon] = v
            continue
        # Cabeceras con variantes de espacio / texto extra
        for rk, rv in raw.items():
            if not rv:
                continue
            if rk == header_key or header_key in rk or rk in header_key:
                out.setdefault(canon, rv)
                break
    return out


OUTPUT_COLUMNS = [
    "nombre",
    "telefono",
    "correo",
    "etapa",
    "estatus",
    "fecha_actualizacion",
    "closer",
    "encargado",
    "recepcion",
    "servicios",
    "cantidad_lineas",
    "valor_contrato",
    "fecha_cierre",
    "red_social",
    "observaciones",
]

# Etiquetas en español para el CSV (Twenty te deja emparejar columna ↔ campo por nombre parecido)
OUTPUT_LABELS_ES = [
    "Nombre empresa",
    "Teléfono",
    "Correo",
    "Etapa venta",
    "Estatus",
    "Última actualización",
    "Closer",
    "Encargado",
    "Recepción / asistente",
    "Servicios activos",
    "Cantidad líneas / servicios",
    "Valor contrato / ARPU",
    "Fecha cierre",
    "Red social / web",
    "Observaciones",
]


def load_workbook(path: Path):
    try:
        import openpyxl
    except ImportError:
        print("Instala openpyxl: python3 -m pip install openpyxl", file=sys.stderr)
        sys.exit(1)
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def extract_rows(wb, sheet_pref: list[str]):
    sheet_name = None
    for cand in sheet_pref:
        if cand in wb.sheetnames:
            sheet_name = cand
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    if sheet_name == "Abril_Funnel" and len(rows) >= 3:
        # fila 0 título, fila 1 cabeceras
        headers_raw = [_norm(c) for c in rows[1]]
        body = rows[2:]
    else:
        headers_raw = [_norm(c) for c in rows[0]]
        body = rows[1:]

    hdr_keys = [_clean_header(h) for h in headers_raw]
    merged_rows: list[dict[str, str]] = []

    for r in body:
        if not any(r):
            continue
        raw = _row_to_dict(headers_raw, r)
        keyed = {_clean_header(k): v for k, v in raw.items()}
        merged_rows.append(_merge_aliases(keyed))

    return sheet_name, merged_rows


def main() -> None:
    p = argparse.ArgumentParser(description="Genera CSV para import Twenty desde reporte xlsx.")
    p.add_argument("xlsx", type=Path, help="Ruta al .xlsx del reporte")
    p.add_argument(
        "--out",
        type=Path,
        default=Path("infra/twenty-import/empresas_seguimiento.csv"),
        help="CSV de salida (UTF-8)",
    )
    args = p.parse_args()

    if not args.xlsx.is_file():
        print(f"No existe archivo: {args.xlsx}", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(args.xlsx)
    try:
        # Abril_Funnel suele tener el embudo completo; Hoja 4 a veces es un subconjunto reciente.
        used_sheet, data = extract_rows(wb, ["Abril_Funnel", "Hoja 4"])
    finally:
        wb.close()

    args.out.parent.mkdir(parents=True, exist_ok=True)
    with args.out.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(OUTPUT_LABELS_ES)
        for m in data:
            if not m.get("nombre"):
                continue
            row = [m.get(c, "") for c in OUTPUT_COLUMNS]
            w.writerow(row)

    print(f"Hoja usada: {used_sheet}")
    print(f"Filas escritas (con nombre): {len([m for m in data if m.get('nombre')])}")
    print(f"CSV generado: {args.out.resolve()}")


if __name__ == "__main__":
    main()
